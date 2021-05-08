# -*- coding: utf-8 -*-
"""
Created on Sat Oct 10 11:55:01 2020

@author: wwr
"""
# from table_extractor import TableExtractor

import os
import pickle
import re
import sys
import traceback
import html
import numpy as np
import openpyxl
import unidecode
from bs4 import BeautifulSoup
from bson.objectid import (ObjectId)
from gensim.models.deprecated import keyedvectors
from scipy import stats
from unidecode import unidecode_expect_nonascii
from dictionary import Dictionary
from log_wp import LogWp
from table import (Attribute, Entity, Link, Table)


class TableExtractorToAlloy(object):
    def __init__(self, xml_path, save_path, config_path):
        self.xml_path = xml_path
        self.save_path = save_path
        self.dict_info = Dictionary(config_path)
        self.list_of_units = self.dict_info.table_units
        self.log_wp = LogWp()

    def get_caption(self, doi, table, format):
        if format == 'html':
            if '10.1016' in doi:
                up = table.parent
                table_root = up.parent
                caption = table_root.find('div', 'caption')
                caption = caption.find('p')
                caption, ref = self._search_for_reference(caption, format)
                caption = unidecode.unidecode(html.unescape(caption.text)).strip()
                return caption, ref
            elif '10.1039' in doi:
                check = table.parent
                check = check.parent
                if check.get('class') == ['rtable__wrapper']:
                    up = table.parent
                    up = up.parent
                    caption = up.previous_sibling
                    if caption is None:
                        return '', []
                    else:
                        caption = caption.previous_sibling
                        if caption is None:
                            return '', []
                        else:
                            caption = caption.find('span')
                            caption, ref = self._search_for_reference(caption, format)
                            caption = unidecode.unidecode(html.unescape(caption.text)).strip()
                            return caption, ref
                else:
                    return '', []
            elif '10.1002' in doi:
                up = table.parent
                caption = up.previous_sibling
                caption = caption.previous_sibling
                if caption is not None:
                    caption.span.decompose()
                    caption, ref = self._search_for_reference(caption, format)
                    caption = unidecode.unidecode(html.unescape(caption.text)).strip()
                    return caption, ref
                else:
                    print('No caption')
                    return '', []
            elif '10.1021' in doi:
                up = table.parent
                if up.get('class') == ['NLM_table-wrap']:
                    caption = up.find('div', 'NLM_caption')
                else:
                    caption = up.previous_sibling
                if caption == ' ':
                    caption = caption.previous_sibling
                if caption is not None:
                    caption, ref = self._search_for_reference(caption, format)
                    caption = unidecode.unidecode(html.unescape(caption.text)).strip()
                    return caption, ref
                else:
                    return '', None
            elif '10.1007' in doi:
                up = table.parent
                caption = up.previous_sibling
                caption = caption.find('p')
                caption, ref = self._search_for_reference(caption, format)
                caption = unidecode.unidecode(html(caption.text)).strip()
                return caption, ref
            else:
                return '', []
        elif format == 'xml':
            if '10.1016' in doi:
                try:
                    caption = table.find('caption')
                    caption, ref = self._search_for_reference(caption, format)
                    caption = unidecode.unidecode(html.unescape(caption.text)).strip()
                except Exception as e:
                    caption = ""
                    ref = ""
                    print(e)
                return caption, ref
            elif '10.1021' in doi:
                caption = table.find('title')
                if caption is None:
                    up = table.parent
                    caption = table.find('title')
                    if caption is None:
                        caption = up.find('caption')
                caption, ref = self._search_for_reference(caption, format)
                caption = unidecode.unidecode(html.unescape(caption.text)).strip()
                return caption, ref
        return '', []

    def get_xml_tables(self, doi, xml):
        all_tables = []
        all_captions = []
        soup = BeautifulSoup(open(xml, 'r+', encoding='utf-8'), 'xml')
        tables = soup.find_all('table')
        if len(tables) == 0:
            soup = BeautifulSoup(open(xml, 'r+', encoding='utf-8'), 'lxml')
            tables = soup.find_all('table-wrap')
        for w, table in enumerate(tables):
            try:
                try:
                    caption, ref = self.get_caption(doi, table, format='xml')
                except Exception as e:
                    print(e)
                    print('Problem in caption')
                    caption = ""
                all_captions.append(caption)
                tab = []
                sup_tab = []
                for t in range(150):
                    tab.append([None] * 150)
                    sup_tab.append([None] * 150)
                rows = table.find_all('row')
                if len(rows) == 0:
                    rows = table.find_all('oasis:row')
                for i, row in enumerate(rows):
                    counter = 0
                    for ent in row:
                        curr_col = 0
                        beg = 0
                        end = 0
                        more_row = 0
                        if type(ent) == type(row):
                            if ent.has_attr('colname'):
                                try:
                                    curr_col = int(ent['colname'])
                                except Exception as e:
                                    print(e)
                                    curr = list(ent['colname'])
                                    for c in curr:
                                        try:
                                            curr_col = int(c)
                                        except Exception as e:
                                            print(e)
                                            continue
                            if ent.has_attr('namest'):
                                try:
                                    beg = int(ent['namest'])
                                except Exception as e:
                                    print(e)
                                    curr = list(ent['namest'])
                                    for c in curr:
                                        try:
                                            beg = int(c)
                                        except Exception as e:
                                            print(e)
                                            continue
                            if ent.has_attr('nameend'):
                                try:
                                    end = int(ent['nameend'])
                                except Exception as e:
                                    print(e)
                                    curr = list(ent['nameend'])
                                    for c in curr:
                                        try:
                                            end = int(c)
                                        except Exception as e:
                                            print(e)
                                            continue
                            if ent.has_attr('morerows'):
                                try:
                                    more_row = int(ent['morerows'])
                                except Exception as e:
                                    print(e)
                                    curr = list(ent['morerows'])
                                    for c in curr:
                                        try:
                                            more_row = int(c)
                                        except Exception as e:
                                            print(e)
                                            continue
                            ent, curr_ref = self._search_for_reference(ent, 'xml')
                            if beg != 0 and end != 0 and more_row != 0:
                                for j in range(beg, end + 1):
                                    for k in range(more_row + 1):
                                        tab[i + k][j - 1] = unidecode.unidecode(
                                            html.unescape(ent.get_text())).strip().replace('\n', ' ')
                                        sup_tab[i + k][j - 1] = curr_ref
                            elif beg != 0 and end != 0:
                                for j in range(beg, end + 1):
                                    tab[i][j - 1] = unidecode.unidecode(
                                        html.unescape(ent.get_text())).strip().replace('\n', ' ')
                                    sup_tab[i][j - 1] = curr_ref
                            elif more_row != 0:
                                for j in range(more_row + 1):
                                    tab[i + j][counter] = unidecode.unidecode(
                                        html.unescape(ent.get_text())).strip().replace('\n', ' ')
                                    sup_tab[i + j][counter] = curr_ref
                            elif curr_col != 0:
                                tab[i][curr_col - 1] = unidecode.unidecode(
                                    html.unescape(ent.get_text())).strip().replace('\n', ' ')
                                sup_tab[i][curr_col - 1] = curr_ref
                            else:
                                counter_ent = counter
                                found = False
                                while not found:
                                    if tab[i][counter_ent] is None:
                                        tab[i][counter_ent] = unidecode.unidecode(
                                            html.unescape(ent.get_text())).strip().replace('\n', ' ')
                                        sup_tab[i][counter_ent] = curr_ref
                                        found = True
                                    else:
                                        counter_ent += 1
                                counter = counter_ent
                            counter = counter + 1 + (end - beg)
                for t, s in zip(tab, sup_tab):
                    for j, k in zip(reversed(t), reversed(s)):
                        if j is None:
                            t.remove(j)
                            s.remove(k)
                for t, s in zip(reversed(tab), reversed(sup_tab)):
                    if len(t) == 0:
                        tab.remove(t)
                        sup_tab.remove(s)
                lens = []
                for t in tab:
                    lens.append(len(t))
                size = stats.mode(lens)[0][0]
                for t, s in zip(tab, sup_tab):
                    if len(t) != size:
                        for j in range(len(t), size):
                            t.append('')
                            s.append([])
                all_tables.append(tab)
            except Exception as e:
                print('Failed to extract XML table')
                table = [[0]]
                print(e)
                all_tables.append(table)
                tb = sys.exc_info()[-1]
                self.log_wp.print_log(traceback.extract_tb(tb, limit=1)[-1][1])
        return all_tables, all_captions

    def get_headers(self, tables, doi):
        all_col_headers = []
        all_row_headers = []
        all_col_indexes = []
        all_row_indexes = []
        for num, table in enumerate(tables):
            try:
                curr = table[0]
                col_index = 0
                for i in range(len(table) - 1):
                    next = table[i + 1]
                    count_curr = 0
                    count_next = 0
                    for cell in curr:
                        try:
                            cell, _ = self.value_extractor(cell)
                        except Exception as e:
                            print(e)
                            if cell != '':
                                count_curr += 1
                    for cell in next:
                        try:
                            cell, _ = self.value_extractor(cell)
                        except Exception as e:
                            print(e)
                            if cell != '':
                                count_next += 1
                    if count_next > count_curr:
                        curr = next
                    else:
                        col_index = 0
                        break
                trans_table = list(map(list, zip(*table)))
                curr_row = trans_table[0]
                row_index = 0
                for i in range(len(trans_table) - 1):
                    next = trans_table[i + 1]
                    count_curr = 0
                    count_next = 0
                    for cell in curr_row:
                        try:
                            cell, _ = self.value_extractor(cell)
                        except Exception as e:
                            print(e)
                            if cell != '':
                                count_curr += 1
                    for cell in next:
                        try:
                            cell, _ = self.value_extractor(cell)
                        except Exception as e:
                            print(e)
                            if cell != '':
                                count_next += 1
                    if count_next > count_curr:
                        curr = next
                    else:
                        row_index = 0
                        break
                row_header = []
                col_header = []
                for i in range(col_index + 1):
                    col_header.extend(table[i])
                for i in range(row_index + 1):
                    row_header.extend(trans_table[i])
                indexes = []
                curr = col_header[0]
                for i in range(len(col_header) - 1):
                    next = col_header[i + 1]
                    if curr == next:
                        indexes.append(i)
                        curr = next
                    else:
                        curr = next
                for i in reversed(indexes):
                    col_header.pop(i)
                indexes = []
                curr = row_header[0]
                for i in range(len(row_header) - 1):
                    next = row_header[i + 1]
                    if curr == next:
                        indexes.append(i)
                        curr = next
                    else:
                        curr = next
                for i in reversed(indexes):
                    row_header.pop(i)
                all_col_headers.append(col_header)
                all_row_headers.append(row_header)
                all_col_indexes.append(col_index)
                all_row_indexes.append(row_index)
            except IndexError as e:
                print("FAILURE: Index self.get_headers table #" + str(num) + " from paper " + str(doi))
                print('IndexError in get headers')
                print(str(e))
                tb = sys.exc_info()[-1]
                self.log_wp.print_log(traceback.extract_tb(tb, limit=1)[-1][1])
        return all_col_headers, all_row_headers, all_col_indexes, all_row_indexes

    def load_embeddings(self, file_loc=None):
        if not file_loc:
            print('Need to specify path to word embedding model')
            print('Materials science training word2vec and fasttext are available for download')
            print('Check the read-me')
        else:
            embeddings = keyedvectors.KeyedVectors.load(file_loc)
            # embeddings.bucket = 2000000
            emb_vocab_ft = dict([('<null>', 0), ('<oov>', 1)] +
                                [(k, v.index + 2) for k, v in embeddings.vocab.items()])
            emb_weights_ft = np.vstack([np.zeros((1, 100)), np.ones((1, 100)), np.array(embeddings.syn0)])

    def _normalize_string(self, string):
        ret_string = ''
        for char in string:
            if re.match(u'[Α-Ωα-ωÅ]', char) is not None:
                ret_string += str(char)
            else:
                ret_string += str(unidecode_expect_nonascii(str(char)))
        return ret_string

    def construct_table_object(self, doi, table, row_ind, col_ind):
        new_table = Table()
        new_table['act_table'] = table
        mat_trans_table = np.array(table).T.tolist()
        mat_table = np.array(table).tolist()
        error_file = []
        try:
            for i, c in enumerate(mat_table[col_ind][(row_ind + 1):]):
                entity = Entity()
                entity['name'] = str(c)
                entity['descriptor'] = str(mat_table[col_ind][row_ind])
                if col_ind > 0:
                    for j in range(col_ind):
                        link = Link()
                        link['name'] = str(mat_table[col_ind - j - 1][i + 1])
                        if link['name'] != entity['name']:
                            attr['links'] = entity['links']
                for j, r in enumerate(mat_trans_table[row_ind][(col_ind + 1):]):
                    attr = Attribute()
                    try:
                        potential_units = unit_regex.search(r).group(0)[1:-1]
                        found_units = [u for u in self.list_of_units if u in potential_units]
                        if len(found_units) > 0:
                            attr['unit'] = unit
                    except Exception as e:
                        print(e)
                        pass
                    attr['name'] = str(r)
                    if row_ind > 0:
                        for k in range(row_ind):
                            link = Link()
                            link['name'] = str(mat_trans_table[row_ind - k - 1][j + 1])
                            if link['name'] != attr['name']:
                                attr['links'].append(link)
                    val, unit = self.value_extractor(str(mat_table[row_ind + j + 1][i + 1]))
                    if type(val) == float:
                        attr['value'] = val
                    else:
                        attr['string_value'] = val
                    if unit is not None:  # overwrites previous unit
                        attr['unit'] = unit
                    entity['attributes'].append(attr)
                new_table['entities'].append(entity)
            return new_table, set(error_file)
        except IndexError as e:
            print("FAILURE: Index construct_table table from paper " + str(doi))
            error_file.append(str(doi))
            print('IndexError in construct object')
            print(str(e))
            return new_table, set(error_file)

    def print_table_object(self, table):
        for ent in table['entities']:
            self.log_wp.print_log('Ent:%s', ent['name'])
            print('Links:')
            for link in ent['links']:
                self.log_wp.print_log(link['name'])
            self.log_wp.print_log('Attr:')
            for att in ent['attributes']:
                self.log_wp.print_log(att['name'])
                self.log_wp.print_log(att['value'])
                for link in att['links']:
                    self.log_wp.print_log(link['name'])
            print('-------')
        print('--------------')

    def value_extractor(self, string):
        original_string = string[:]
        extracted_unit = None
        balance_syn = ['balance', 'bal', 'bal.', 'other.', 'other']
        if string.lower() in balance_syn:
            return 'balance', extracted_unit
        units = [u for u in self.list_of_units if u in string]
        if units:
            extracted_unit = max(units)
            string = string.replace(extracted_unit, '')
        # e.g. already in int or float form: 12.5 -> 12.5
        try:
            return float(string), extracted_unit
        except Exception as e:
            print(e)
            pass
        # e.g. 12.5 - 13.5 -> 13.0
        range_regex = re.compile('\d+\.?\d*\s*-\s*\d+\.?\d*')
        try:
            ranges = range_regex.search(string).group().split('-')
            average = (float(ranges[0]) + float(ranges[1])) / 2.0
            return average, extracted_unit
        except Exception as e:
            print(e)
            pass
        # e.g. 12.2 (5.2) -> 12.2
        bracket_regex = re.compile('(\d+\.?\d*)\s*\(\d*.?\d*\)')
        try:
            extracted_value = float(bracket_regex.search(string).group(1))
            return float(extracted_value), extracted_unit
        except Exception as e:
            print(e)
            pass
        # e.g. 12.3 ± 0.5 -> 12.3
        plusmin_regex = re.compile('(\d+\.?\d*)(\s*[±+-]+\s*\d+\.?\d*)')
        try:
            extracted_value = float(plusmin_regex.search(string).group(1))
            return extracted_value, extracted_unit
        except AttributeError:
            pass
        # e.g. <0.05 -> 0.05  |  >72.0 -> 72.0    | ~12 -> 12
        lessthan_roughly_regex = re.compile('([<]|[~]|[>])=?\s*\d+\.*\d*')
        try:
            extracted_value = lessthan_roughly_regex.search(string).group()
            num_regex = re.compile('\d+\.*\d*')
            extracted_value = num_regex.search(extracted_value).group()
            return float(extracted_value), extracted_unit
        except Exception as e:
            print(e)
            pass
        # e.g. 0.4:0.6 (ratios)
        if ':' in string:
            split = string.split(":")
            try:
                extracted_value = round(float(split[0]) / float(split[1]), 3)
                return extracted_value, extracted_unit
            except Exception as e:
                print(e)
                pass
        return original_string, None

    def load_composition_elements(self, domain=None):
        # Compositional elements to help in correclty identifiying the orientation of tables in specific domains
        if domain == 'geopolymers':
            material_constituents = ['Al2O3', 'SiO2']
            constituent_threshold = 2
            remaining = None
        elif domain == 'steel':
            material_constituents = ['Fe', 'Cr', 'Cu', 'C', 'Ti', 'Ni', 'Mo', 'Mn']
            constituent_threshold = 4
            remaining = ['Fe']
        elif domain == 'titanium':
            material_constituents = ['Ti', 'Fe', 'C']
            constituent_threshold = 2
            remaining = ['Fe']
        elif domain == 'zeolites':
            material_constituents = (
                ['Si/Ge', 'DMAP/T', 'HF/T', 'H2O/T', '(Si + Ge)/Al', 'SiO2', 'GeO2', 'SDA', 'HF', 'H2O', 'Ge', 'Si',
                 'SiO2/Al2O3', 'Si/Al',
                 'R(OH)2/Si', 'F-/Si', '(Si + Ge)/Zr', 'Al', 'SDA/Si', 'H2O/Si', 'OH/Si', 'Si/H2O', 'Si/OH', 'Ge/Si',
                 'Si/Ti',
                 'MeO',
                 'SiO2/GeO2', 'TMHDA', 'TMEDA', 'TEOS', 'NH4F', 'Al/T', 'N,N-Diethylethylenediamine', 'NaGaGeO4',
                 'NaGaO2',
                 'Na2GeO3*H2O',
                 'SOD', 'NaNO2', 'NaOH'])
            constituent_threshold = 2
            remaining = None
        elif domain == 'aluminum':
            material_constituents = ['Al', 'Cu', 'Mn', 'Si', 'O', 'Mg']
            constituent_threshold = 2
            remaining = None
        elif domain == 'alloys':
            material_constituents = ['Ag', 'Al', 'Ar', 'As', 'Au', 'B', 'Ba', 'Be', 'Bi', 'Br', 'C', 'Ca', 'Cd', 'Ce',
                                     'Cl', 'Co', 'Cr', 'Cs', 'Cu', 'Dy',
                                     'Er', 'Eu', 'F', 'Fe', 'Ga', 'Gd', 'Ge', 'H', 'Hf', 'Hg', 'Ho', 'I', 'In', 'Ir',
                                     'K', 'La', 'Li', 'Lu', 'Md', 'Mg',
                                     'Mn', 'Mo', 'N', 'Na', 'Nb', 'Nd', 'Ni', 'O', 'Os', 'P', 'Pb', 'Pd', 'Pr', 'Pt',
                                     'Rb', 'Re', 'Rh', 'Ru', 'S', 'Sb',
                                     'Sc', 'Se', 'Si', 'Sm', 'Sn', 'Sr', 'Ta', 'Tb', 'Te', 'Th', 'Ti', 'Tl', 'Tm', 'U',
                                     'V', 'W', 'Y', 'Yb', 'Zn', 'Zr']
            constituent_threshold = 2
            remaining = ['Fe', 'Al', 'Ti']

    def set_balance(self, entity, balance_pos, cumsum):
        if cumsum < 1:
            entity['attributes'][balance_pos]['value'] = 1.0 - cumsum
        else:
            entity['attributes'][balance_pos]['value'] = 100.0 - cumsum

    def get_links(self, entity):
        list_of_names = []
        for attr in entity['attributes']:
            list_of_names.append(attr['name'])
        if len(set(list_of_names)) < 3:
            for attr in entity['attributes']:
                if len(attr['links']) > 0:
                    swapped = attr['name']
                    attr['name'] = attr['links'][0]['name']
                    attr['links'][0]['name'] = swapped

    def check_if_balanced(self, cumsum):
        if cumsum > 1:
            if 100 - cumsum < 1.5:
                return True
            else:
                return False
        else:
            if 1 - cumsum < 0.015:
                return True
            else:
                return False

    def _search_for_reference(self, soup, format):
        if format == 'html':
            ref = soup.find_all('a')
            tags = []
            if len(ref) == 0:
                text = soup.text
                refs = re.findall('\[\D\]', text)
                if len(refs) == 0:
                    return soup, tags
                else:
                    text = re.split('\[\D\]', text)
                    text = ''.join(text)
                    soup.string = text
                    return soup, refs
            else:
                for r in ref:
                    tag = soup.a.extract()
                    tags.append(tag.text)
                return soup, tags
        elif format == 'xml':
            ref = soup.find_all('xref')
            tags = []
            if len(ref) == 0:
                if soup.name == 'caption':
                    return soup, tags
                ref = soup.find_all('sup')
                for r in ref:
                    text = r.text.split(',')
                    for t in text:
                        if len(t) == 1 and t.isalpha():
                            tags.append(t)
                            soup.sup.decompose()
                return soup, tags
            else:
                for r in ref:
                    if len(r.text) < 4:
                        tag = soup.xref.extract()
                        tags.append(tag.text)
                return soup, tags


def get_extraction_outcome(xml_path, save_path, config_path):
    table_extractor_m = TableExtractorToAlloy(xml_path, save_path, config_path)
    all_error_file = []
    xml_name = os.listdir(xml_path)
    log_wp = LogWp()
    for file_i in range(len(os.listdir(xml_path))):
        tables = None
        all_tables = []
        doi = xml_name[file_i].replace(".xml", "")  # 根据doi来选择
        doi = doi.replace("-", "/", 1)
        xml_n = xml_name[file_i]  # 根据doi来选择
        file = xml_path + '/' + str(xml_n)
        try:
            tables, captions = table_extractor_m.get_xml_tables(doi, file)
        except Exception as e:
            print(e)
            all_error_file.append(doi)
            tables = None
            captions = None
        if tables:
            cols, rows, col_inds, row_inds = table_extractor_m.get_headers(tables, doi)
            tab = []
            for table, row_ind, col_ind in zip(tables, row_inds, col_inds):
                curr, error_file = (table_extractor_m.construct_table_object(doi, table, row_ind, col_ind))
                if curr:
                    tab.append(curr)
                if error_file:
                    all_error_file.append(str(doi))
            for i, (t, caption) in enumerate(zip(tab, captions)):
                if t is not None:
                    t['order'] = i
                    t['_id'] = ObjectId()
                    t['caption'] = caption
                    t['paper_doi'] = doi
                    all_tables.append(t)
                    log_wp.print_log('Success: Extracted Tables from %s', doi)
            xls = openpyxl.Workbook()
            sheet_id = 1
            if all_tables:
                for table in all_tables:
                    sht_new = xls.create_sheet(str(sheet_id))
                    act_table = table['act_table']
                    caption = table['caption']
                    row_len = len(act_table[0])
                    doi = table['paper_doi']
                    sht_new.cell(1, 1, str(doi))
                    sht_new.cell(2, 1, str(caption))
                    start_row = 3
                    for row in act_table:
                        len_row = len(row)
                        for index in range(len_row):
                            sht_new.cell(start_row, index + 1, row[index])
                        start_row += 1
                    sheet_id += 1
                del xls['Sheet']
                xls.save(save_path + '/' + str(file_i) + ".xlsx")
    return all_error_file, len(all_error_file)

