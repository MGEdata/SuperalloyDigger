# -*- coding: utf-8 -*-
"""
Created on Sat Oct  3 21:45:27 2020

@author: win
"""

import xlrd
import os
from dictionary import Dictionary
import re
import nltk
import openpyxl
from log_wp import Log_wp

class Table_extraction:
    def __init__(self, excels_path, C_path, prop_name='solvus'):
        self.excels_path = excels_path
        self.C_path = C_path
        self.prop_name = prop_name
        self.dict_info = Dictionary(self.C_path)
        self.ele_list = self.dict_info.ele_list
        self.e_pattern = self.dict_info.table_e_pattern
        self.ratio_pattern = self.dict_info.table_ratio_pattern
        self.prop_pattern = self.dict_info.table_prop_pattern
        self.unit_pattern = self.dict_info.unit_pattern_table
        self.number_pattern = self.dict_info.table_number_pattern
        self.ele_to_abr = self.dict_info.ele_to_abr
        self.prop_pattern_words = self.dict_info.table_prop_pattern_words
        self.log_wp = Log_wp()

    def composition_triple_extraction(self):
        file_list = os.listdir(self.excels_path)
        composition_all = {}
        for excel_path in file_list:
            try:
                file = xlrd.open_workbook(self.excels_path + '/' + excel_path)
                all_material = []               
                for sheet_i in range(len(file.sheets())):
                    try:
                        sheet = file.sheet_by_index(sheet_i)
                        topic = sheet.row_values(1)[0] 
                        if 'composition' in topic.lower():
                            target_ele_row = []
                            target_ele_col = []
                            search_outcome = []
                            ele_loc = None
                            for line_index in range(2,len(sheet.col_values(0))):
                                search_line = sheet.row_values(line_index)
                                unit_i = 0
                                for unit in search_line:
                                    outcome = re.findall(self.e_pattern,str(unit))
                                    if outcome and str(unit) in self.ele_list:
                                        target_ele_row.append(line_index)
                                        target_ele_col.append(unit_i)
                                        search_outcome.append(unit)
                                    unit_i += 1  
                                if search_outcome:
                                    ele_loc = line_index
                                    break
                            if ele_loc:
                                first_col = sheet.col_values(0)
                                dict_info = Dictionary(self.C_path)
                                alloy_replace = dict_info.table_alloy_to_replace
                                alloy_common_type = dict_info.alloy_writing_type
                                alloy_blank_type = dict_info.alloy_blank_type
                                for alloy_model,replace in alloy_replace.items():
                                    alloy_part = re.findall(alloy_model,str(topic))
                                    for alloy in alloy_part:
                                        find_part = re.findall(replace[0],str(alloy))
                                        alloy_out = alloy.replace(find_part[0],replace[1])
                                        topic = topic.replace(alloy,alloy_out)
                                    outcome_name = []
                                    topic_tokenize = nltk.word_tokenize(topic)
                                    for word in topic_tokenize:
                                        for pattern_1 in alloy_common_type:
                                            outcome_common = re.findall(pattern_1,str(word))
                                            if outcome_common:
                                                outcome_name.append(word)
                                                break        
                                    for pattern_2 in alloy_blank_type:
                                        outcome_blank = re.findall(pattern_2,str(topic))
                                        if outcome_blank and outcome_blank[0] not in outcome_name:
                                            outcome_name.append(outcome_blank[0])
                                            break
                                len_col = len(sheet.row_values(3))
                                alloy_name_col = None
                                alloy_name_search = []
                                if len_col <= 3:
                                    for col_i in range(len_col):
                                        col_info = sheet.col_values(col_i)
                                        if col_i == 0:
                                            col_info = sheet.col_values(col_i)[2:]
                                        if col_info:
                                            for cell in col_info:
                                                for pattern_1 in alloy_common_type:
                                                    outcome_common = re.findall(pattern_1,str(cell))
                                                    if outcome_common:
                                                        alloy_name_col = col_i
                                                        alloy_name_search.append(col_i)
                                                for pattern_2 in alloy_blank_type:
                                                    outcome_blank = re.findall(pattern_2,str(cell))
                                                    if outcome_blank:
                                                        alloy_name_col = col_i
                                                        alloy_name_search.append(col_i)
                                else:
                                    for col_i in range(3):
                                        col_info = sheet.col_values(col_i)
                                        if col_i == 0:
                                            col_info = sheet.col_values(col_i)[2:]
                                        if col_info:
                                            for cell in col_info:
                                                for pattern_1 in alloy_common_type:
                                                    outcome_common = re.findall(pattern_1,str(cell))
                                                    if outcome_common:
                                                        alloy_name_col = col_i
                                                        alloy_name_search.append(col_i)
                                                for pattern_2 in alloy_blank_type:
                                                    outcome_blank = re.findall(pattern_2,str(cell))
                                                    if outcome_blank:
                                                        alloy_name_col = col_i
                                                        alloy_name_search.append(col_i)
                                if not alloy_name_search:
                                    alloy_name_col = 0
                                else:
                                    alloy_name_col = alloy_name_search[0]
                                first_col = sheet.col_values(0)
                                ele_first = []
                                for unit in first_col:
                                    firstcol_search = re.findall(self.e_pattern,str(unit))
                                    if firstcol_search:
                                        ele_first.append(unit)
                                if len(ele_first) <= 2:
                                    if len(first_col) > 4:
                                        e_search = re.findall(self.e_pattern,str(sheet.col_values(0)[ele_loc]))
                                        if e_search and outcome_name and len(outcome_name) == 1:
                                            for index_row in range(ele_loc+1,len(first_col)):
                                                composition_single = {}
                                                composition_single['material'] = outcome_name[0].replace('~',' ')
                                                composition_single['doi']=first_col[0]
                                                ratio_find_topic = re.findall(self.ratio_pattern,str(topic))
                                                ratio_find_col = re.findall(self.ratio_pattern,str(first_col[index_row]))
                                                for table_head in sheet.row_values(2):
                                                    ratio_find_head = re.findall(self.ratio_pattern,str(table_head))
                                                    if ratio_find_head:
                                                        composition_single['percentage'] = ratio_find_head[0]
                                                        break
                                                if ratio_find_topic:
                                                    composition_single['percentage'] = ratio_find_topic[0]
                                                elif ratio_find_col:
                                                    composition_single['percentage'] = ratio_find_col[0]   
                                                for ele_index in range(len(sheet.row_values(2))):
                                                    ele_name = sheet.row_values(ele_loc)[ele_index]
                                                    if ele_name in tuple(self.ele_to_abr.keys()):
                                                        ele_name = self.ele_to_abr[ele_name]
                                                    number = sheet.row_values(index_row)[ele_index]
                                                    composition_single[ele_name]=number
                                                all_material.append(composition_single)
                                        if not e_search:
                                            for index_row in range(ele_loc+1,len(first_col)):
                                                if first_col[index_row]:
                                                    composition_single = {}
                                                    name_col = sheet.col_values(alloy_name_col)
                                                    if outcome_name and len(outcome_name) == 1 and not alloy_name_search:
                                                        composition_single['material'] = outcome_name[0].replace('~', ' ')
                                                    else:
                                                        composition_single['material'] = name_col[index_row]
                                                    composition_single['doi']=first_col[0]
                                                    ratio_find_topic = re.findall(self.ratio_pattern,str(topic))
                                                    ratio_find_col = re.findall(self.ratio_pattern,str(first_col[index_row]))
                                                    for table_head in sheet.row_values(2):
                                                        ratio_find_head = re.findall(self.ratio_pattern,str(table_head))
                                                        if ratio_find_head:
                                                            composition_single['percentage'] = ratio_find_head[0]
                                                            break
                                                    if ratio_find_topic:
                                                        composition_single['percentage'] = ratio_find_topic[0]
                                                    elif ratio_find_col:
                                                        composition_single['percentage'] = ratio_find_col[0]
                                                    ratio_find_unit = re.findall(self.ratio_pattern,str(first_col[index_row]))
                                                    if ratio_find_unit:
                                                        composition_single['percentage'] = ratio_find_unit[0]
                                                    for ele_index in range(len(sheet.row_values(ele_loc)[1:])):
                                                        ele_name = sheet.row_values(ele_loc)[1:][ele_index]
                                                        if ele_name in tuple(self.ele_to_abr.keys()):
                                                            ele_name = self.ele_to_abr[ele_name]
                                                        number = sheet.row_values(index_row)[ele_index+1]
                                                        composition_single[ele_name]=number
                                                    all_material.append(composition_single)
                                    else:
                                        composition_single = {}
                                        first_col_1 = sheet.row_values(3)[0]
                                        e_search = re.findall(self.e_pattern,str(sheet.col_values(0)[ele_loc]))
                                        ratio_find_col = re.findall(self.ratio_pattern,str(first_col_1))
                                        for table_head in sheet.row_values(2):
                                            ratio_find_head = re.findall(self.ratio_pattern,str(table_head))
                                            if ratio_find_head:
                                                composition_single['percentage'] = ratio_find_head[0]
                                                break
                                        if ratio_find_col:
                                            composition_single['percentage'] = ratio_find_col[0]
                                        ratio_find_topic = re.findall(self.ratio_pattern,str(topic))
                                        if ratio_find_topic:
                                            composition_single['percentage'] = ratio_find_topic[0]
                                        if outcome_name and e_search:
                                            composition_single['material'] = outcome_name[0].replace('~',' ')
                                            composition_single['doi']=first_col[0]
                                            for ele_index in range(len(sheet.row_values(2))):
                                                ele_name = sheet.row_values(ele_loc)[ele_index]
                                                number = sheet.row_values(3)[ele_index]
                                                if ele_name in tuple(self.ele_to_abr.keys()):
                                                    ele_name = self.ele_to_abr[ele_name]
                                                composition_single[ele_name]=number
                                            all_material.append(composition_single)
                                        elif outcome_name and not e_search:
                                            if len(outcome_name) == 1:
                                                composition_single['material'] = outcome_name[0].replace('~',' ')
                                            else:
                                                composition_single['material'] = sheet.row_values(ele_loc+1)[alloy_name_col]
                                            composition_single['doi']=first_col[0]
                                            for ele_index in range(len(sheet.row_values(2)[1:])):
                                                ele_name = sheet.row_values(ele_loc)[1:][ele_index]
                                                number = sheet.row_values(3)[1:][ele_index]
                                                if ele_name in tuple(self.ele_to_abr.keys()):
                                                    ele_name = self.ele_to_abr[ele_name]
                                                composition_single[ele_name]=number
                                            all_material.append(composition_single)
                                        elif not outcome_name and not e_search:
                                            composition_single['material'] = sheet.row_values(ele_loc+1)[alloy_name_col]
                                            composition_single['doi']=first_col[0]
                                            m_name = sheet.row_values(ele_loc)[0]
                                            composition_single[m_name] = first_col[3]
                                            for ele_index in range(len(sheet.row_values(2)[1:])):
                                                ele_name = sheet.row_values(ele_loc)[1:][ele_index]
                                                number = sheet.row_values(3)[1:][ele_index]
                                                if ele_name in tuple(self.ele_to_abr.keys()):
                                                    ele_name = self.ele_to_abr[ele_name]
                                                composition_single[ele_name]=number
                                            all_material.append(composition_single)                                                                                                          
                                        elif not outcome_name and e_search:
                                            composition_single['material'] = None
                                            composition_single['doi']=first_col[0]
                                            for ele_index in range(len(sheet.row_values(2))):
                                                ele_name = sheet.row_values(ele_loc)[ele_index]
                                                number = sheet.row_values(3)[ele_index]
                                                if ele_name in tuple(self.ele_to_abr.keys()):
                                                    ele_name = self.ele_to_abr[ele_name]
                                                composition_single[ele_name]=number
                                            all_material.append(composition_single)
                                else:
                                    ele_row = sheet.row_values(ele_loc-1)
                                    len_elerow = len(ele_row)
                                    for index_col in range(1,len_elerow):
                                        if ele_row[index_col] :
                                            composition_single = {}
                                            if outcome_name and len(outcome_name) == 1 and len_elerow <=2:
                                                material_name = outcome_name[0].replace('~',' ')
                                            else:
                                                material_name = ele_row[index_col]
                                            composition_single['material']=material_name
                                            composition_single['doi']=first_col[0]
                                            ratio_find_topic = re.findall(self.ratio_pattern,str(topic))
                                            ratio_find_col = re.findall(self.ratio_pattern,str(material_name))
                                            if ratio_find_topic:
                                                composition_single['percentage'] = ratio_find_topic[0]
                                            elif ratio_find_col:
                                                composition_single['percentage'] = ratio_find_col[0]
                                            for ele_index in range(len(sheet.col_values(0)[ele_loc:])):
                                                ele_name = sheet.col_values(0)[ele_loc:][ele_index]
                                                number = sheet.col_values(index_col)[ele_loc+ele_index]
                                                if ele_name in tuple(self.ele_to_abr.keys()):
                                                    ele_name = self.ele_to_abr[ele_name]
                                                composition_single[ele_name]=number
                                            all_material.append(composition_single)
                        if all_material:
                            break
                    except Exception as e:
                        self.log_wp.print_log(str(e))
                        self.log_wp.print_log("An error in the %s of %s!",sheet_i,excel_path)
                if all_material:
                    composition_all[excel_path] = all_material
            except Exception as e:
                self.log_wp.print_log("can't open this file, name of file is %s", str(excel_path))
                self.log_wp.print_log("Error is %s", str(e))
                self.log_wp.print_log("--"*25)
        return composition_all

    def property_info_extraction(self):
        file_list = os.listdir(self.excels_path)
        property_all = {}
        number_prop = 0
        K_path = []
        for excel_path in file_list:
           try:
                file = xlrd.open_workbook(self.excels_path + '/' + excel_path)
                all_material = []
                for sheet_i in range(len(file.sheets())):
                   try:
                        sheet = file.sheet_by_index(sheet_i)
                        topic = sheet.row_values(1)[0]
                        search_outcome = []
                        target_prop_row = []
                        target_prop_col = []
                        for line_index in range(2,len(sheet.col_values(0))):
                            search_line = sheet.row_values(line_index)[1:]
                            unit_i = 1
                            for unit in search_line:
                                outcome_words = None
                                for pattern in self.prop_pattern[self.prop_name]:
                                    outcome = re.findall(pattern,str(unit))
                                    if all(word in str(unit) for word in self.prop_pattern_words[self.prop_name]):
                                        outcome_words = unit
                                    if outcome:
                                        break
                                if outcome or outcome_words:
                                    target_prop_row.append(line_index)
                                    
                                    target_prop_col.append(unit_i)
                                    search_outcome.append(unit)
                                unit_i += 1
                        if any(search_outcome):
                            first_col = sheet.col_values(0)
                            alloy_replace = Dictionary(self.C_path).table_alloy_to_replace
                            for alloy_model,replace in alloy_replace.items():
                                alloy_part = re.findall(alloy_model,str(topic))
                                for alloy in alloy_part:
                                    find_part = re.findall(replace[0],str(alloy))
                                    alloy_out = alloy.replace(find_part[0],replace[1])
                                    topic = topic.replace(alloy,alloy_out)
                            alloy_common_type = Dictionary(self.C_path).alloy_writing_type
                            alloy_blank_type = Dictionary(self.C_path).alloy_blank_type
                            outcome_name = []
                            topic_tokenize = nltk.word_tokenize(topic)
                            for word in topic_tokenize:
                                for pattern_1 in alloy_common_type:
                                    outcome_common = re.findall(pattern_1,str(word))
                                    if outcome_common:
                                        outcome_name.append(word)
                                        break        
                            for pattern_2 in alloy_blank_type:
                                outcome_blank = re.findall(pattern_2,str(topic))
                                if outcome_blank:
                                    outcome_name.append(outcome_blank[0])
                                    break
                            fc_ns = []
                            for cell in sheet.col_values(0)[1:]:
                                fc_n = re.findall(self.number_pattern[self.prop_name],str(cell))
                                alphabet_search = re.findall("[A-Za-z]",str(cell))
                                if fc_n and not alphabet_search:
                                    fc_ns.append(cell)
                            len_col = len(sheet.row_values(3))
                            alloy_name_col = None
                            alloy_name_search = []
                            if len_col <= 3:
                                for col_i in range(len_col):
                                    col_info = sheet.col_values(col_i)
                                    if col_i == 0:
                                        col_info = sheet.col_values(col_i)[2:]
                                    if col_info:
                                        for cell in col_info:
                                            for pattern_1 in alloy_common_type:
                                                outcome_common = re.findall(pattern_1,str(cell))
                                                if outcome_common:
                                                    alloy_name_col = col_i
                                                    alloy_name_search.append(col_i)
                                            for pattern_2 in alloy_blank_type:
                                                outcome_blank = re.findall(pattern_2,str(cell))
                                                if outcome_blank:
                                                    alloy_name_col = col_i
                                                    alloy_name_search.append(col_i)
                            else:
                                for col_i in range(3):
                                    col_info = sheet.col_values(col_i)
                                    if col_i == 0:
                                        col_info = sheet.col_values(col_i)[2:]
                                    if col_info:
                                        for cell in col_info:
                                            for pattern_1 in alloy_common_type:
                                                outcome_common = re.findall(pattern_1,str(cell))
                                                if outcome_common:
                                                    alloy_name_col = col_i
                                                    alloy_name_search.append(col_i)
                                            for pattern_2 in alloy_blank_type:
                                                outcome_blank = re.findall(pattern_2,str(cell))
                                                if outcome_blank:
                                                    alloy_name_col = col_i
                                                    alloy_name_search.append(col_i)     
                            if not alloy_name_search:
                                alloy_name_col = 0
                            else:
                                alloy_name_col = alloy_name_search[0]
                            if len(first_col) > 4:
                                for prop_i in range(len(target_prop_row)): 
                                    sub_label = []
                                    curr_col = []
                                    for index_row in range(target_prop_row[prop_i]+1,len(first_col)):
                                        unit_search_parts = []
                                        unit_search_parts.append(topic)
                                        if len(fc_ns) == 0:
                                            name_col = sheet.col_values(alloy_name_col)
                                            material_name = name_col[index_row]
                                            property_single = {}
                                            property_single['table_topic'] = first_col[1]
                                            number = sheet.row_values(index_row)[target_prop_col[prop_i]]
                                            number_inspect = re.findall(self.number_pattern[self.prop_name],str(number))
                                            prop_name = sheet.row_values(target_prop_row[prop_i])[target_prop_col[prop_i]]
                                            unit_search_parts.append(first_col[index_row])
                                            unit_search_parts.append(number)
                                            for unit in sheet.row_values(target_prop_row[0]):
                                                unit_search_parts.append(unit)
                                            for row_s in range(2,target_prop_row[prop_i]+1):
                                                unit_search_parts.append(sheet.row_values(row_s)[target_prop_col[prop_i]])
                                            if number_inspect:
                                                one_info = {}
                                                for prop_index in range(len(sheet.row_values(target_prop_row[prop_i]))):
                                                    prop_name_line = sheet.row_values(target_prop_row[prop_i])[prop_index]
                                                    number_line_line = sheet.row_values(index_row)[prop_index]
                                                    one_info[prop_name_line]=number_line_line
                                                curr_col.append(number)
                                                property_single[prop_name]=number
                                                property_single['other_info']=one_info
                                                property_single['material']=material_name
                                                property_single['doi'] = first_col[0]
                                                if sub_label:
                                                    property_single['child_tag'] = sub_label
                                                for item in unit_search_parts:
                                                    unit_find = re.findall(self.unit_pattern[self.prop_name],str(item))
                                                    if unit_find:
                                                        property_single['unit'] = unit_find[0].replace('degC','°C')
                                                        K_path.append(excel_path)
                                                if 'unit' not in property_single.keys():
                                                    property_single['unit'] = 'no mentioned'   
                                            elif not number_inspect and len(curr_col) != 0:
                                                property_single['material']=material_name
                                                property_single['doi'] = first_col[0]
                                                property_single[prop_name]= number
                                                if sub_label:
                                                    property_single['child_tag'] = sub_label
                                                for item in unit_search_parts:
                                                    unit_find = re.findall(self.unit_pattern[self.prop_name],str(item))
                                                    if unit_find:
                                                        property_single['unit'] = unit_find[0].replace('degC','°C')
                                                        K_path.append(excel_path)
                                                        break
                                                if 'unit' not in property_single.keys():
                                                    property_single['unit'] = 'no mentioned'   
                                            elif not number_inspect and len(curr_col) == 0:
                                                if number and not property_single:
                                                    if number != '-' and number != '--':
                                                        sub_label.append(number)
                                            if property_single:
                                                all_material.append(property_single)
                                        if first_col[index_row] and len(fc_ns) != 0 and len(outcome_name)==1:
                                            material_name = outcome_name[0].replace('~',' ')
                                            property_single = {}
                                            property_single['table_topic'] = first_col[1]
                                            unit_search_parts.append(first_col[index_row])
                                            for row_s in range(2,target_prop_row[prop_i]+1):
                                                unit_search_parts.append(sheet.row_values(row_s)[target_prop_col[prop_i]])
                                            prop_name = sheet.row_values(target_prop_row[prop_i])[target_prop_col[prop_i]]
                                            number = sheet.row_values(index_row)[target_prop_col[prop_i]]
                                            number_inspect = re.findall(self.number_pattern[self.prop_name],str(number))
                                            unit_search_parts.append(number)
                                            if number_inspect:
                                                property_single[prop_name]=number
                                                property_single['material']=material_name
                                                property_single['doi'] = first_col[0]
                                                for item in unit_search_parts:
                                                    unit_find = re.findall(self.unit_pattern[self.prop_name],str(item))
                                                    if unit_find:
                                                        property_single['unit'] = unit_find[0].replace('degC','°C')
                                                        K_path.append(excel_path)
                                                        break
                                                if 'unit' not in property_single.keys():
                                                    property_single['unit'] = 'no mentioned'   
                                            elif not number_inspect and len(curr_col) != 0:
                                                property_single['material']=material_name
                                                property_single['doi'] = first_col[0]
                                                property_single[prop_name]=number
                                                if sub_label:
                                                    property_single['child_tag'] = sub_label
                                                for item in unit_search_parts:
                                                    unit_find = re.findall(self.unit_pattern[self.prop_name],str(item))
                                                    if unit_find:
                                                        property_single['unit'] = unit_find[0].replace('degC','°C')
                                                        K_path.append(excel_path)
                                                        break
                                                if 'unit' not in property_single.keys():
                                                    property_single['unit'] = 'no mentioned'                                                  
                                            elif not number_inspect and len(curr_col) == 0:
                                                if number and not property_single:
                                                    sub_label.append(number)
                                            if property_single:
                                                all_material.append(property_single)
                            else:
                                unit_search_parts = []
                                property_single = {}
                                property_single['table_topic'] = first_col[1]
                                alloy_replace = Dictionary(self.C_path).table_alloy_to_replace
                                for alloy_model,replace in alloy_replace.items():
                                    alloy_part = re.findall(alloy_model,str(topic))
                                    for alloy in alloy_part:
                                        find_part = re.findall(replace[0],str(alloy))
                                        alloy_out = alloy.replace(find_part[0],replace[1])
                                        topic = topic.replace(alloy,alloy_out)
                                alloy_common_type = Dictionary(self.C_path).alloy_writing_type
                                alloy_blank_type = Dictionary(self.C_path).alloy_blank_type
                                outcome_name = []
                                topic_tokenize = nltk.word_tokenize(topic)
                                for word in topic_tokenize:
                                    for pattern_1 in alloy_common_type:
                                        outcome_common = re.findall(pattern_1,str(word))
                                        if outcome_common:
                                            outcome_name.append(word)
                                            break        
                                for pattern_2 in alloy_blank_type:
                                    outcome_blank = re.findall(pattern_2,str(topic))
                                    if outcome_blank and outcome_blank[0] not in outcome_name:
                                        outcome_name.append(outcome_blank[0])
                                        break
                                unit_search_parts.append(first_col[3])
                                unit_search_parts.append(topic)
                                for row_s in range(2,4):
                                    for prop_i in range(len(target_prop_row)):
                                        unit_search_parts.append(sheet.row_values(row_s)[target_prop_col[prop_i]])
                                number_search = re.findall(self.number_pattern[self.prop_name],str(sheet.col_values(0)[2]))
                                if outcome_name and number_search:
                                    for prop_i in range(len(target_prop_row)):
                                        property_single['material'] = outcome_name[0].replace('~',' ')
                                        property_single['doi'] = first_col[0]
                                        number = sheet.row_values(3)[target_prop_col[prop_i]]
                                        unit_search_parts.append(number)
                                        for item in unit_search_parts:
                                            unit_find = re.findall(self.unit_pattern[self.prop_name],str(item))
                                            if unit_find:
                                                property_single['unit'] = unit_find[0].replace('degC','°C')
                                                K_path.append(excel_path)
                                                break
                                        if 'unit' not in property_single.keys():
                                            property_single['unit'] = 'no mentioned'
                                        prop_name = sheet.row_values(target_prop_row[prop_i])[target_prop_col[prop_i]]
                                        property_single[prop_name]=number
                                        all_material.append(property_single)
                                elif not outcome_name and not number_search:
                                    for prop_i in range(len(target_prop_row)):
                                        property_single[sheet.col_values(2)[0]] = first_col[3]
                                        property_single['doi'] = first_col[0]
                                        number = sheet.row_values(3)[target_prop_col[prop_i]]
                                        unit_search_parts.append(number)
                                        for item in unit_search_parts:
                                            unit_find = re.findall(self.unit_pattern[self.prop_name],str(item))
                                            if unit_find:
                                                property_single['unit'] = unit_find[0].replace('degC','°C')
                                                K_path.append(excel_path)
                                                break
                                        if 'unit' not in property_single.keys():
                                            property_single['unit'] = 'no mentioned'
                                        prop_name = sheet.row_values(target_prop_row[prop_i])[target_prop_col[prop_i]]
                                        property_single[prop_name]=number
                                        all_material.append(property_single)  
                                elif not outcome_name and number_search:
                                    for prop_i in range(len(target_prop_row)):
                                        property_single['material'] = 'no mentioned'
                                        property_single['doi'] = first_col[0]
                                        number = sheet.row_values(3)[target_prop_col[prop_i]]
                                        unit_search_parts.append(number)
                                        for item in unit_search_parts:
                                            unit_find = re.findall(self.unit_pattern,str(item))
                                            if unit_find:
                                                property_single['unit'] = unit_find[0].replace('degC','°C')
                                                K_path.append(excel_path)
                                                break
                                        if 'unit' not in property_single.keys():
                                            property_single['unit'] = 'no mentioned'
                                        prop_name = sheet.row_values(target_prop_row[prop_i])[target_prop_col[prop_i]]
                                        property_single[prop_name]=number
                                        all_material.append(property_single)
                   except Exception as e:
                       self.log_wp.print_log("An error in file: %s-sheet:%s---%s!", excel_path, sheet_i, e)
                if all_material:
                    number_prop += 1
                    property_all[excel_path] = all_material
           except Exception as e:
               self.log_wp.print_log("can't open %s ", excel_path)
               self.log_wp.print_log(str(e))
               self.log_wp.print_log("--"*25)
        return property_all


class Get_targetinfo():
    def __init__(self,all_info,out_path):
        self.related_allinfo = all_info
        self.out_path = out_path
        self.log_wp = Log_wp()

    def structure_ele(self):
        xls = openpyxl.Workbook()
        sht = xls.create_sheet(index=0)
        sht.cell(1,1,"File_name")
        sht.cell(1,2,"DOIs")
        sht.cell(1,3,"Material")
        sht.cell(1,4,"Percentage")
        sht.cell(1,5,"Element_info and other_info")
        start_row = 2
        for file,ele in self.related_allinfo.items():
            for material in ele:
                sht.cell(start_row,1,file)
                if 'doi' in material.keys():
                    sht.cell(start_row,2,material['doi'])
                    material.pop('doi')
                if 'material' in material.keys():
                    material_name = material['material']
                    noisy = re.findall('\s*\[.+\]',str(material_name))
                    if noisy:
                        for puc in noisy:
                            material_name = str(material_name).replace(puc,'')
                    sht.cell(start_row,3,material_name)
                    material.pop('material')
                if 'percentage' in material.keys():
                    sht.cell(start_row,4,material['percentage'])
                    material.pop('percentage')
                if material:
                    sht.cell(start_row,5,str(material))
                start_row += 1
        self.log_wp.excel_save(xls, self.out_path)
    
    def structure_prop(self,prop_pattern,prop_name_s):
        xls = openpyxl.Workbook()
        sht = xls.create_sheet(0)
        sht.cell(1,1,"File_name")
        sht.cell(1,2,"DOIs")
        sht.cell(1,3,"Table_topic")
        sht.cell(1,4,"Material")
        sht.cell(1,5,"unit")
        sht.cell(1,6,"Property_name")
        sht.cell(1,7,"Property_value")
        sht.cell(1,8,"Child_tag")
        sht.cell(1,9,"Other_info")
        start_row = 2
        for file,ele in self.related_allinfo.items():
            for material in ele:
                get_prop = None
                sht.cell(start_row,1,file)
                if 'doi' in material.keys():
                    sht.cell(start_row,2,material['doi'])
                    material.pop('doi')
                if 'table_topic' in material.keys():
                    sht.cell(start_row, 3, material['table_topic'])
                    material.pop('table_topic')
                if 'material' in material.keys():
                    material_name = material['material']
                    noisy = re.findall('\s*\[.+\]',str(material_name))
                    if noisy:
                        for puc in noisy:
                            material_name = str(material_name).replace(puc,'')
                    sht.cell(start_row,4,material_name)
                    material.pop('material')
                if 'unit' in material.keys():
                    unit_replace = material['unit'].replace('degC','°C')
                    sht.cell(start_row,5,unit_replace)
                    material.pop('unit')
                if 'child_tag' in material.keys():
                    sht.cell(start_row,8,str(material['child_tag']))
                    material.pop('child_tag')
                if 'other_info' in material.keys():
                    sht.cell(start_row,9,str(material['other_info']))
                    material.pop('other_info')
                if len(material) == 1:
                    for prop_name,value in material.items():
                        sht.cell(start_row,6,str(prop_name))
                        sht.cell(start_row,7,str(value))
                elif len(material) >= 1:
                    for prop_name,value in material.items():
                        for pattern in prop_pattern[prop_name_s]:
                            prop_search = re.findall(pattern,str(prop_name))
                            if prop_search:
                                sht.cell(start_row,6,str(prop_name))
                                sht.cell(start_row,7,str(value))
                                get_prop = True
                                break
                        if get_prop:
                            break
                start_row += 1
        self.log_wp.excel_save(xls, self.out_path)

def transform_comp_outcome(all_composition,ele_list):
    gather_outcome = []
    for file_name,t_info in all_composition.items():
        sole_info = {}
        all_triple_info = []
        for sole_m_info in t_info:
            triple_info = {}
            sole_material = ''
            sole_doi = sole_m_info['doi']
            sole_m_info.pop('doi')
            if 'material' in sole_m_info.keys():
                sole_material = sole_m_info['material']
                noisy = re.findall('\s*\[.+\]',str(sole_material))
                if noisy:
                    for puc in noisy:
                        sole_material = str(sole_material).replace(puc,'')
                triple_info['material'] = sole_material
                sole_m_info.pop('material')            
            for element in ele_list:
                if element in sole_m_info.keys():
                    triple_info[element] = sole_m_info[element]
                    sole_m_info.pop(element)
            if sole_m_info:
                triple_info["other_eleinfo"] = sole_m_info
            all_triple_info.append(triple_info)
        sole_info[sole_doi] = all_triple_info
        gather_outcome.append(sole_info)
        
    return gather_outcome
