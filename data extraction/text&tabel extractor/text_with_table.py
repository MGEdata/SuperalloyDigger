# -*- coding: utf-8 -*-
"""
Created on Wed Oct 28 17:00:53 2020

@author: win
"""

import re
import openpyxl
import os
from get_full_text import Filter_text
from pre_processor import Pre_processor
from sentence_positioner import Sentence_Positioner
from T_pre_processor import T_pre_processor
from Phrase_parse import Phrase_parse
from Relation_extraciton import Relation_extraciton
from output_modified_triple import Table_extraction,Get_targetinfo
import copy
from dictionary import Dictionary
from log_wp import Log_wp
import nltk

property_list = ['solvus','density','liquidus','solidus']

class Acquire_all_target_info():
    def __init__(self, c_path, origin_text_path, prop_list, excels_path, out_path):
        self.c_path = c_path
        self.prop_list = prop_list
        self.origin_text_path = origin_text_path
        self.excels_path = excels_path
        self.dict_info = Dictionary(self.c_path)
        self.out_path = out_path
        self.log_wp = Log_wp()

    def mkdir(self,file_name):
        pathd=os.getcwd()+'\\'+file_name
        if os.path.exists(pathd):
            for root, dirs, files in os.walk(pathd, topdown=False):
                for name in files:
                    os.remove(os.path.join(root, name))
                for name in dirs:
                    os.rmdir(os.path.join(root, name))
            os.rmdir(pathd)
        os.mkdir(pathd)

    def get_doi_fromtxt(self,txt_path):
        text_name = txt_path.replace(".txt","")
        doi = text_name.replace("-","/",1)
        return doi

    def get_abrre(self,text,prop_name):
        processor = T_pre_processor(text, prop_name, self.c_path)
        text = processor.processor()
        sentences = nltk.sent_tokenize(text)
        sentences_split = text.split(" ")
        alloy_write_type = self.dict_info.alloy_writing_type
        len_type = len(alloy_write_type)
        abbre_to_alloy = {}
        for sent in sentences:
            processor = T_pre_processor(sent, prop_name, self.c_path)
            filter_data = processor.processor()
            words =  nltk.word_tokenize(filter_data)
            for word in words:
                abbre_exist = None
                for type_i in range(0, len_type):
                    outcome = re.findall(alloy_write_type[type_i], word)
                    outcome_alloy = None
                    if outcome:
                        abbre = "("+word+")"
                        if abbre in sentences_split:
                            index_alloy = sentences_split.index(abbre) - 1
                            alloy = sentences_split[index_alloy]
                            for type_j in range(0, len_type):
                                outcome_alloy = re.findall(alloy_write_type[type_j], alloy)
                                if outcome_alloy:
                                    abbre_to_alloy[word] = alloy
                                    abbre_exist = True
                                    break
                    if outcome_alloy:
                        break
        return abbre_to_alloy

    def get_text_triple(self,prop_name):
        self.mkdir('output_tt')
        text_path = r"output_tt\full_text"
        self.mkdir(text_path)
        FT = Filter_text(self.origin_text_path,text_path)
        txt_name = FT.process()
        length = len(os.listdir(self.origin_text_path))
        all_txt_info = []
        for i in range(0,length):
            n_path = os.listdir(self.origin_text_path)[i]
            doi = self.get_doi_fromtxt(n_path)
            file = open(text_path + '/' + n_path,'r',encoding='utf-8')
            data = file.read()
            pre_processor = Pre_processor(data,self.c_path)
            filter_txt = pre_processor.pre_processor()
            file_origin = open(self.origin_text_path + '/' + n_path,'r',encoding='utf-8')
            data_origin = file_origin.read()
            abbre_pairs = self.get_abrre(data_origin,prop_name)
            positioner = Sentence_Positioner(filter_txt,prop_name,self.c_path)
            target_sents = positioner.target_sent()
            for index,sent in target_sents.items():
                processor = T_pre_processor(sent,prop_name,self.c_path)
                filter_data = processor.processor()
                parse = Phrase_parse(filter_data,prop_name,self.c_path)
                sub_order,sub_id,object_list = parse.alloy_sub_search()
                RE = Relation_extraciton(prop_name,filter_data,sub_order,sub_id,object_list,self.c_path,abbre_pairs)
                all_outcome = RE.triple_extraction()
                if all_outcome:
                    for id_m,info in all_outcome.items():
                        sole_info = {}
                        sole_info['doi'] = doi
                        sole_info['material'] = info[0]
                        sole_info['prop_name'] = info[1]
                        sole_info['prop_value'] = info[2]
                        all_txt_info.append(sole_info)
        return all_txt_info

    def gather_tableinfo_textinfo(self,all_txt_info,table_info,prop_name,prop_pattern,unit_pattern_text):
        gather_outcome = []
        for file_name,t_info in table_info.items():
            sole_info = {}
            all_triple_info = []
            sole_doi = []
            for sole_m_info in t_info:
                triple_info = {}
                sole_material = ''
                sole_unit = ''
                sole_propname = ''
                sole_value = ''
                triple_info['source'] = 'table'
                if 'doi' in sole_m_info.keys():
                    plus_doi = sole_m_info['doi']
                    sole_doi.append(plus_doi)
                    sole_m_info.pop('doi')
                if 'material' in sole_m_info.keys():
                    sole_material = sole_m_info['material']
                    noisy = re.findall('\s*\[.+\]',str(sole_material))
                    if noisy:
                        for puc in noisy:
                            sole_material = str(sole_material).replace(puc,'')
                    triple_info['material'] = sole_material
                    sole_m_info.pop('material')
                if 'unit' in sole_m_info.keys():
                    sole_unit = sole_m_info['unit']
                    triple_info['unit'] = sole_unit
                    sole_m_info.pop('unit')
                if 'other_info' in sole_m_info.keys():
                    sole_other_info = sole_m_info['other_info']
                    triple_info['other_prop_info'] = sole_other_info
                    sole_m_info.pop('other_info')
                if 'child_tag' in sole_m_info.keys():
                    sole_tag_info = sole_m_info['child_tag']
                    triple_info['child_tag'] = sole_tag_info
                    sole_m_info.pop('child_tag')
                if 'table_topic' in sole_m_info.keys():
                    sole_tag_info = sole_m_info['table_topic']
                    triple_info['table_topic'] = sole_tag_info
                    sole_m_info.pop('table_topic')
                if len(sole_m_info) == 1:
                    for prop_name_t,value in sole_m_info.items():
                        sole_propname = str(prop_name_t)
                        triple_info['prop_name'] = sole_propname
                        sole_value = str(value)
                        triple_info['value'] = sole_value
                elif len(sole_m_info) >= 1:
                    get_prop = None
                    for prop_name_t,value in sole_m_info.items():
                        for pattern in prop_pattern[prop_name]:
                            prop_search = re.findall(pattern,str(prop_name_t))
                            if prop_search:
                                sole_propname = str(prop_name_t)
                                triple_info['prop_name'] = sole_propname
                                sole_value = str(value)
                                triple_info['value'] = sole_value
                                get_prop = True
                                break
                        if get_prop:
                            break
                all_triple_info.append(triple_info)
            if list(set(sole_doi)):
                sole_info[list(set(sole_doi))[0]] = all_triple_info
                gather_outcome.append(sole_info)
        gather = 0
        for q in gather_outcome:
            k = tuple(q.keys())[0]
            i = q[k]
            for n in i:
                for w,v in n.items():
                    if w =='value':
                        gather += 1
        self.log_wp.print_log("gather number :%s", gather)
        copy_all_txt_info = copy.copy(all_txt_info)
        if copy_all_txt_info:
            all_text = 0
            all_gather_doi = []
            for info_one in gather_outcome:
                all_gather_doi.append(tuple(info_one.keys())[0])
            for triple_info_sole in copy_all_txt_info:
                if triple_info_sole['doi'] in all_gather_doi:
                    all_text += 1
                    plus_info = {}
                    plus_info['source'] = 'text'
                    plus_info['prop_name'] = triple_info_sole['prop_name']
                    prop_value = triple_info_sole['prop_value']
                    plus_info['material'] = triple_info_sole['material']
                    unit_search = re.findall(unit_pattern_text[prop_name],str(prop_value))
                    if unit_search:
                        plus_info['unit'] = unit_search[0]
                        prop_value = prop_value.replace(unit_search[0],'')
                        plus_info['value'] = prop_value
                    else:
                        plus_info['unit'] = ""
                        plus_info['value'] = prop_value
                    for get_info in gather_outcome:
                        if tuple(get_info.keys())[0] == triple_info_sole['doi']:
                            get_info[triple_info_sole['doi']].append(plus_info)
                if triple_info_sole['doi'] not in all_gather_doi:
                    all_text += 1
                    plus_info = {}
                    full_info = {}
                    sole_triple = []
                    plus_info['source'] = 'text'
                    plus_info['prop_name'] = triple_info_sole['prop_name']
                    prop_value = triple_info_sole['prop_value']
                    plus_info['material'] = triple_info_sole['material']
                    unit_search = re.findall(unit_pattern_text[prop_name],str(prop_value))
                    if unit_search:
                        plus_info['unit'] = unit_search[0]
                        prop_value = prop_value.replace(unit_search[0],'')
                        plus_info['value'] = prop_value
                    else:
                        plus_info['unit'] = ""
                        plus_info['value'] = prop_value
                    if plus_info:
                        sole_triple.append(plus_info)
                        full_info[triple_info_sole['doi']] = sole_triple
                        gather_outcome.append(full_info)
                        all_gather_doi.append(triple_info_sole['doi'])
            self.log_wp.print_log("all_text number :%s", all_text)
        return gather_outcome

    def transform_comp_outcome(self,all_composition):
        ele_list = self.dict_info.ele_list
        gather_outcome = []
        for file_name,t_info in all_composition.items():
            sole_info = {}
            all_triple_info = []
            for sole_m_info in t_info:
                triple_info = {}
                sole_material = ''
                sole_doi = sole_m_info['doi']
                sole_m_info.pop('doi')
                if 'material' in sole_m_info.keys():
                    sole_material = sole_m_info['material']
                    noisy = re.findall('\[.+\]',str(sole_material))
                    if noisy:
                        for puc in noisy:
                            sole_material = str(sole_material).replace(puc,'')
                    triple_info['material'] = sole_material
                    sole_m_info.pop('material')
                for element in ele_list:
                    if element in sole_m_info.keys():
                        triple_info[element] = sole_m_info[element]
                        sole_m_info.pop(element)
                if sole_m_info:
                    triple_info["other_eleinfo"] = sole_m_info
                all_triple_info.append(triple_info)
            sole_info[sole_doi] = all_triple_info
            gather_outcome.append(sole_info)
        return gather_outcome

    def allinfo_dependencyparse(self,comp_info,prop_info):
        all_ele_doi = []
        all_prop_doi = []
        outcome = []
        for doi_info_ele in comp_info:
            ele_doi = tuple(doi_info_ele.keys())[0]
            all_ele_doi.append(ele_doi)
        for doi_info_prop in prop_info:
            prop_doi = tuple(doi_info_prop.keys())[0]
            all_prop_doi.append(prop_doi)
        prop_info_modified = copy.copy(prop_info)
        for doi_info_ele in comp_info:
            ele_doi = tuple(doi_info_ele.keys())[0]
            if ele_doi in all_prop_doi:
                for doi_info_prop in prop_info:
                    prop_doi = tuple(doi_info_prop.keys())[0]
                    plus_info = {}
                    all_doi_info = []
                    if ele_doi == prop_doi:
                        if doi_info_prop in prop_info_modified:
                            prop_info_modified.remove(doi_info_prop)
                        ele_doi_fullinfo = doi_info_ele[ele_doi]
                        ele_allname = []
                        prop_allname = []
                        pop_name = []
                        for one_material_ele in ele_doi_fullinfo:
                            if 'material' in one_material_ele.keys():
                                ele_m_name = one_material_ele['material']
                                ele_allname.append(ele_m_name)
                        modified_ele_allname = []
                        for name in ele_allname:
                            space_search = re.findall('\s',str(name))
                            if space_search:
                                name_list = str(name).split()
                                modified_ele_allname.append(str(name))
                                for name_sepe in name_list:
                                    modified_ele_allname.append(name_sepe)
                            else:
                                modified_ele_allname.append(name)
                        for one_material_prop in doi_info_prop[prop_doi]:
                            if 'material' in one_material_prop.keys():
                                prop_m_name = one_material_prop['material']
                                prop_allname.append(prop_m_name)
                                if prop_m_name not in modified_ele_allname and len(ele_doi_fullinfo) == 1:
                                    if one_material_prop['source'] == 'table':
                                        combine_info = {}
                                        for prop_name,prop_value in one_material_prop.items():
                                            combine_info[prop_name] = prop_value
                                        for ele_name,ele_value in ele_doi_fullinfo[0].items():
                                            combine_info[ele_name] = ele_value
                                        all_doi_info.append(combine_info)
                                    else:
                                        all_doi_info.append(one_material_prop)
                                if prop_m_name not in modified_ele_allname and len(ele_doi_fullinfo) != 1:
                                    all_doi_info.append(one_material_prop)
                                if prop_m_name in modified_ele_allname:
                                    for one_material_ele in ele_doi_fullinfo:
                                       if 'material' in one_material_ele.keys():
                                           ele_m_name = one_material_ele['material']
                                           space_search = re.findall('\s',str(ele_m_name))
                                           if space_search:
                                               ele_m_name_split = ele_m_name.split()
                                               if prop_m_name in ele_m_name_split or prop_m_name == ele_m_name:
                                                   pop_name.append(ele_m_name)
                                                   combine_info = {}
                                                   for prop_name,prop_value in one_material_prop.items():
                                                       combine_info[prop_name] = prop_value
                                                   for ele_name,ele_value in one_material_ele.items():
                                                       combine_info[ele_name] = ele_value
                                                   all_doi_info.append(combine_info)
                                           else:
                                               if prop_m_name == ele_m_name :
                                                   combine_info = {}
                                                   for prop_name,prop_value in one_material_prop.items():
                                                       combine_info[prop_name] = prop_value
                                                   for ele_name,ele_value in one_material_ele.items():
                                                       combine_info[ele_name] = ele_value
                                                   all_doi_info.append(combine_info)
                        for one_material_ele in ele_doi_fullinfo:
                            if 'material' in one_material_ele.keys():
                                ele_m_name = one_material_ele['material']
                                if ele_m_name not in pop_name:
                                    if ele_m_name not in prop_allname:
                                         all_doi_info.append(one_material_ele)
                    if all_doi_info:
                        plus_info[ele_doi] = all_doi_info
                        outcome.append(plus_info)
            else:
                outcome.append(doi_info_ele)
        for extra_prop in prop_info_modified:
            outcome.append(extra_prop)
        return outcome

    def structureinfo_toexcel(self,all_structureinfo,out_path):
        ele_list = self.dict_info.ele_list
        xls = openpyxl.Workbook()
        sht = xls.create_sheet(0)
        sht = xls.create_sheet(index=0)
        sht.cell(1,1,"Source")
        sht.cell(1,2,"DOIs")
        sht.cell(1,3,"table_topic")
        sht.cell(1,4,"material")
        sht.cell(1,5,"Property_name")
        sht.cell(1,6,"Property_value")
        sht.cell(1,7,"Unit")
        col_n = 8
        row_now = 2
        sht.cell(1,col_n,str("other_element_info"))
        col_n += 1
        sht.cell(1,col_n,str("other_property_info"))
        col_n += 1
        sht.cell(1,col_n,str("child_tag"))
        col_n += 1
        for ele in ele_list:
            sht.cell(1,col_n,ele)
            col_n += 1
        for m_info in all_structureinfo:
            doi = tuple(m_info.keys())[0]
            length_m_info = m_info[doi]
            for index_m in range(len(length_m_info)):
                sht.cell(row_now,2,doi)
                material_now = length_m_info[index_m]
                if 'source' in material_now.keys():
                    sht.cell(row_now,1,str(material_now['source']))
                if 'table_topic' in material_now.keys():
                    sht.cell(row_now,3,str(material_now['table_topic']))
                if 'material' in material_now.keys():
                    sht.cell(row_now,4,str(material_now['material']))
                if 'prop_name' in material_now.keys():
                    sht.cell(row_now,5,str(material_now['prop_name']))
                if 'value' in material_now.keys():
                    sht.cell(row_now,6,str(material_now['value']))
                if 'unit' in material_now.keys():
                    sht.cell(row_now,7,str(material_now['unit']))
                if "other_eleinfo" in material_now.keys():
                    sht.cell(row_now,8,str(material_now['other_eleinfo']))
                if "other_prop_info" in material_now.keys():
                    sht.cell(row_now,9,str(material_now['other_prop_info']))
                if "child_tag" in material_now.keys():
                    sht.cell(row_now,10,str(material_now["child_tag"]))
                col_ele = 11
                for ele in ele_list:
                    if ele in material_now.keys():
                        sht.cell(row_now,col_ele,material_now[ele])
                    col_ele += 1
                row_now += 1
        del xls['Sheet']
        self.log_wp.excel_save(xls, out_path)

    def run(self):
        prop_pattern = self.dict_info.table_prop_pattern
        number_pattern = self.dict_info.table_number_pattern
        unit_pattern_table = self.dict_info.unit_pattern_table
        unit_pattern_text = self.dict_info.table_unit_pattern_text
        for prop_name in self.prop_list:
            self.mkdir('output_tt')
            text_path = r"output_tt\full_text"
            self.mkdir(text_path)
            all_txt_info = self.get_text_triple( prop_name)
            target_property = prop_name  # 'density' 'liquidus'  'solidus'  'solvus'
            te = Table_extraction(self.excels_path, self.c_path, prop_name=target_property)
            info_all = te.property_info_extraction()
            i_l = 0
            for k, v in info_all.items():
                i_l += len(v)
            all_composition = te.composition_triple_extraction()
            gather_outcome = self.gather_tableinfo_textinfo(all_txt_info, info_all, prop_name, prop_pattern,
                                                       unit_pattern_text)
            gather = 0
            for q in gather_outcome:
                k = tuple(q.keys())[0]
                i = q[k]
                gather += len(i)
            ele_transform = self.transform_comp_outcome(all_composition)
            all_structureinfo = self.allinfo_dependencyparse(ele_transform,
                                                        gather_outcome)
            b = 0
            for a in all_structureinfo:
                k = tuple(a.keys())[0]
                i = a[k]
                for n in i:
                    for w, v in n.items():
                        if w == 'value':
                            b += 1
            out_path = self.out_path +'/' + str(prop_name) + '.xlsx'
            self.structureinfo_toexcel(all_structureinfo,out_path)

